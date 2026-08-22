"""Workbook to SQLite, once, at build time.

The application never parses a spreadsheet. `scripts/build_db.py` produces
`data/parcelpilot.db`, which is committed, so the hosted app has nothing to do
at startup and any reviewer gets byte-identical data.

Two rules govern the conversion:

**Local time is Asia/Kolkata.** The workbook stores naive datetimes and states
its timezone once, in the README sheet. Every timestamp is localised on the way
in and written with an offset, because a naive value resolves differently on
every host and would shift every cancellation window silently.

**Absent stays absent.** Blank cells become NULL, never 0 and never "". A
missing pickup time is a different fact from a pickup at the epoch.
"""

from __future__ import annotations

import sqlite3
from collections.abc import Iterator
from datetime import datetime
from pathlib import Path
from typing import Any, Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

REPO_ROOT: Final = Path(__file__).resolve().parents[2]
WORKBOOK_PATH: Final = REPO_ROOT / "data" / "raw" / "ParcelPilot_Assessment_Data.xlsx"
SCHEMA_PATH: Final = Path(__file__).with_name("schema.sql")

#: Stated once in the workbook README: "2026-08-16 11:00 Asia/Kolkata".
WORKBOOK_TZ: Final = ZoneInfo("Asia/Kolkata")

ACCOUNT_COLUMNS: Final = (
    "account_id",
    "account_name",
    "plan",
    "status",
    "csm",
    "contract_file",
    "premium_support",
    "notes",
)
ORDER_COLUMNS: Final = (
    "order_id",
    "account_id",
    "carrier",
    "status",
    "booked_at",
    "pickup_window_start",
    "pickup_window_end",
    "pickup_actual_at",
    "shipment_fee_inr",
    "carrier_fault",
    "customer_fault",
    "cancellation_requested_at",
    "notes",
)
TICKET_COLUMNS: Final = (
    "ticket_id",
    "account_id",
    "created_at",
    "status",
    "subject",
    "description",
    "channel",
    "assigned_to",
    "last_customer_message_at",
    "historical_resolution",
)

_TIMESTAMP_COLUMNS: Final = frozenset(
    {
        "booked_at",
        "pickup_window_start",
        "pickup_window_end",
        "pickup_actual_at",
        "cancellation_requested_at",
        "created_at",
        "last_customer_message_at",
    }
)
_BOOLEAN_COLUMNS: Final = frozenset({"premium_support", "carrier_fault", "customer_fault"})


class EtlError(RuntimeError):
    """The workbook does not look the way the loader expects."""


def build_database(workbook: Path | str = WORKBOOK_PATH, target: Path | str | None = None) -> Path:
    """Rebuild the structured store from the workbook. Idempotent."""
    from src.config import get_settings

    workbook_path = Path(workbook)
    target_path = Path(target) if target else get_settings().db_path
    if not workbook_path.is_file():
        raise EtlError(f"workbook not found: {workbook_path}")

    sheets = _read_sheets(workbook_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(target_path)
    try:
        # The schema drops before it creates, so a rebuild over an existing
        # file replaces rather than appends.
        conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
        conn.executemany(_insert("meta", ("key", "value")), sorted(_meta_rows(sheets["README"])))
        _load(conn, "accounts", ACCOUNT_COLUMNS, sheets["accounts"])
        _load(conn, "orders", ORDER_COLUMNS, sheets["orders"])
        _load(conn, "tickets", TICKET_COLUMNS, sheets["tickets"])
        conn.execute("PRAGMA foreign_keys = ON")
        _assert_referential_integrity(conn)
        conn.commit()
    finally:
        conn.close()
    return target_path


def _read_sheets(path: Path) -> dict[str, list[tuple[Any, ...]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        missing = {"README", "accounts", "orders", "tickets"} - set(workbook.sheetnames)
        if missing:
            raise EtlError(f"workbook is missing sheet(s): {sorted(missing)}")
        return {
            name: list(workbook[name].iter_rows(values_only=True)) for name in workbook.sheetnames
        }
    finally:
        workbook.close()


def _meta_rows(readme: list[tuple[Any, ...]]) -> Iterator[tuple[str, str]]:
    """Extract build provenance from the README sheet.

    `as_of` is the single most consequential value in the pack, so it is read
    from the workbook rather than configured, and stored so a test can assert
    the configured value matches it.
    """
    labels = {
        "dataset snapshot": "as_of",
        "currency": "currency",
    }
    seen: set[str] = set()
    for row in readme:
        if len(row) < 2 or row[0] is None or row[1] is None:
            continue
        key = labels.get(str(row[0]).strip().lower())
        if key is None:
            continue
        value = str(row[1]).strip()
        if key == "as_of":
            value = _parse_snapshot_cell(value).isoformat()
        seen.add(key)
        yield key, value

    if "as_of" not in seen:
        raise EtlError("README sheet has no 'Dataset snapshot' row; AS_OF cannot be derived")


def _parse_snapshot_cell(raw: str) -> datetime:
    """Parse `2026-08-16 11:00 Asia/Kolkata` from the README sheet."""
    head, _, zone_name = raw.rpartition(" ")
    try:
        return datetime.fromisoformat(head).replace(tzinfo=ZoneInfo(zone_name))
    except Exception as exc:
        raise EtlError(f"cannot parse dataset snapshot {raw!r} from the README sheet") from exc


def _load(
    conn: sqlite3.Connection, table: str, columns: tuple[str, ...], sheet: list[tuple[Any, ...]]
) -> None:
    if not sheet:
        raise EtlError(f"sheet for {table} is empty")

    header = [str(c).strip() if c is not None else "" for c in sheet[0]]
    missing = set(columns) - set(header)
    if missing:
        raise EtlError(f"{table} sheet is missing column(s): {sorted(missing)}")
    index = {name: position for position, name in enumerate(header)}

    rows = [
        tuple(_coerce(column, row[index[column]]) for column in columns)
        for row in sheet[1:]
        if any(cell is not None for cell in row)
    ]
    # One multi-row insert, not one statement per row.
    conn.executemany(_insert(table, columns), rows)


def _insert(table: str, columns: tuple[str, ...]) -> str:
    placeholders = ", ".join("?" * len(columns))
    return f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})"


def _coerce(column: str, value: Any) -> Any:
    """Convert one cell, preserving absence.

    Blank cells return None so the column stays NULL. Booleans become 0/1
    because SQLite has no boolean type. Timestamps are localised and rendered
    with an offset.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if column in _BOOLEAN_COLUMNS:
        return (
            int(bool(value)) if not isinstance(value, str) else int(value.strip().lower() == "true")
        )
    if column in _TIMESTAMP_COLUMNS:
        return _to_ist(value, column).isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def _to_ist(value: Any, column: str) -> datetime:
    if isinstance(value, datetime):
        moment = value
    elif isinstance(value, str):
        try:
            moment = datetime.fromisoformat(value.strip())
        except ValueError as exc:
            raise EtlError(f"cannot parse {column}={value!r} as a datetime") from exc
    else:
        raise EtlError(f"{column} holds {type(value).__name__}, expected a datetime")

    # The workbook stores wall time in Asia/Kolkata and says so once, in the
    # README. Attaching it here is what stops the offset from being guessed
    # later, one host at a time.
    return moment.replace(tzinfo=WORKBOOK_TZ) if moment.tzinfo is None else moment


def _assert_referential_integrity(conn: sqlite3.Connection) -> None:
    violations = conn.execute("PRAGMA foreign_key_check").fetchall()
    if violations:
        raise EtlError(f"foreign key violations after load: {violations}")
